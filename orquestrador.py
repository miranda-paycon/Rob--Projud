import os
import openpyxl
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from paycon import extracao_elaw
from projudi_ba import regra_captcha, busca_processo, peticionamento, download_comprovante

PASTA_RELATORIOS = os.path.join(os.path.dirname(__file__), "temp", "relatorios")
os.makedirs(PASTA_RELATORIOS, exist_ok=True)


def ler_processos_do_excel(caminho_excel):
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    cabecalhos = [cell.value for cell in ws[1]]
    idx_processo = cabecalhos.index("Processo")
    processos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        valor = row[idx_processo]
        if valor:
            processos.append(str(valor).strip())
    return processos


def gerar_relatorio_pdf(falhas):
    if not falhas:
        print("Nenhuma falha registrada, relatório não gerado.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = os.path.join(PASTA_RELATORIOS, f"relatorio_falhas_{timestamp}.pdf")

    doc = SimpleDocTemplate(nome, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Relatório de Falhas — Automação PROJUDI", styles["Title"]))
    elementos.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f"Total de falhas: {len(falhas)}", styles["Normal"]))
    elementos.append(Spacer(1, 12))

    dados = [["#", "Processo", "Motivo"]]
    for i, (processo, motivo) in enumerate(falhas, start=1):
        dados.append([str(i), processo, motivo])

    tabela = Table(dados, colWidths=[30, 200, 280])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    print(f"Relatório gerado: {nome}")
    return nome


def main():
    print("=== INICIANDO ORQUESTRADOR ===")
    falhas = []

    try:
        # Etapa 1: Extrair planilha do eLaw
        print("\n[1] Extraindo tarefas do eLaw...")
        caminho_excel = extracao_elaw.executar()
        print(f"Planilha gerada: {caminho_excel}")

        # Etapa 2: Ler processos da planilha
        print("\n[2] Lendo processos da planilha...")
        processos = ler_processos_do_excel(caminho_excel)
        print(f"{len(processos)} processos encontrados.")

        # Etapa 3: Login no PROJUDI
        print("\n[3] Realizando login no PROJUDI...")
        context, page, playwright = regra_captcha.executar()

        # Etapa 4: Para cada processo, buscar no PROJUDI
        total = len(processos)
        for i, numero_processo in enumerate(processos, start=1):
            print(f"\n[4] Processando {i}/{total}: {numero_processo}")
            try:
                encontrou = busca_processo.executar(context, page, numero_processo)

                if encontrou:
                    print(f"Processo {numero_processo} encontrado — aguardando próximos passos...")
                    # Etapa 5: Peticionamento (quando implementado)
                    # peticionamento.executar(context, page, dados)
                    # Etapa 6: Download (quando implementado)
                    # download_comprovante.executar(context, page)
                    break
                else:
                    falhas.append((numero_processo, "Sem botão Peticionar"))

            except Exception as e:
                print(f"Erro ao processar {numero_processo}: {e}")
                falhas.append((numero_processo, f"Erro no captcha / falha: {str(e)[:80]}"))
                try:
                    busca_processo.voltar_para_busca(page)
                except:
                    pass

        print("\n=== TODOS OS PROCESSOS CONCLUÍDOS ===")

    except Exception as e:
        print(f"\nErro geral no orquestrador: {e}")
        falhas.append(("GERAL", f"Erro inesperado: {str(e)[:80]}"))

    finally:
        gerar_relatorio_pdf(falhas)
        input("Pressione ENTER para fechar o navegador...")
        try:
            context.close()
            playwright.stop()
        except:
            pass


if __name__ == "__main__":
    main()
